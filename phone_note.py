# -*- coding: utf-8; py-compile-optimize: 1 -*-
import random
import pandas as pd
import numpy as np
from faker import Faker
from datetime import datetime, timedelta
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk, messagebox
import sys
import os

class ExcelEditor:
    def __init__(self, master):
        self.master = master
        self.data = {}  # 存储所有工作表数据 {sheet_name: DataFrame}
        self.treeviews = {}  # 存储Treeview组件
        
        # 初始化界面
        self.create_ui()
        self.load_data()

    def create_ui(self):
        """创建编辑器界面"""
        # 工具栏
        toolbar = ttk.Frame(self.master)
        toolbar.pack(fill=tk.X, pady=5)
        
        ttk.Button(toolbar, text="添加事件", command=self.show_add_dialog).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar, text="删除选中", command=self.delete_selected).pack(side=tk.LEFT, padx=5)
        ttk.Button(toolbar, text="保存修改", command=self.save_data).pack(side=tk.RIGHT, padx=5)
        # 操作提示
        tip_frame = ttk.Frame(self.master)
        tip_frame.pack(fill=tk.X, pady=3)
        
        # 使用不同样式突出提示
        style = ttk.Style()
        style.configure("Tip.TLabel", foreground="#666", font=('微软雅黑', 9))
        
        tip_label = ttk.Label(
            tip_frame,
            text="操作提示：双击表格修改数据，修改后点击【保存修改】按钮",
            style="Tip.TLabel"
        )
        tip_label.pack(side=tk.LEFT, padx=10)

        # 页签容器
        self.notebook = ttk.Notebook(self.master)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)

    def load_data(self):
        """加载并排序数据"""
        try:
            with pd.ExcelFile(resource_path('events.xlsx')) as xls:
                for sheet_name in xls.sheet_names:
                    df = xls.parse(sheet_name)
                    # 按权重降序排序
                    self.data[sheet_name] = df.sort_values(by='权重', ascending=False)
                    self.create_sheet_tab(sheet_name)
        except Exception as e:
            messagebox.showerror("错误", f"读取文件失败：{str(e)}")

    def create_sheet_tab(self, sheet_name):
        """创建排序后的页签"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text=sheet_name)
        
        tree = ttk.Treeview(frame, columns=("事件", "权重"), show='headings')
        tree.heading("事件", text="事件")
        tree.heading("权重", text="权重", command=lambda: self.sort_column(tree, "权重", False))
        
        # 配置列...
        
        # 插入排序后的数据
        for _, row in self.data[sheet_name].iterrows():
            tree.insert("", tk.END, values=tuple(row))
        
        self.treeviews[sheet_name] = tree

    def sort_column(self, tree, col, reverse):
        """动态排序列"""
        data = [(tree.set(child, col), child) for child in tree.get_children('')]
        data.sort(reverse=reverse, key=lambda x: float(x[0]) if col == "权重" else x[0])
        
        for index, (_, child) in enumerate(data):
            tree.move(child, '', index)
        
        tree.heading(col, command=lambda: self.sort_column(tree, col, not reverse))

    def update_sheet_data(self, sheet_name, tree):
        """更新数据并保持排序"""
        items = tree.get_children()
        data = [tree.item(item, "values") for item in items]
        
        # 转换为DataFrame并排序
        df = pd.DataFrame(data, columns=["事件", "权重"])
        df["权重"] = df["权重"].astype(float)
        self.data[sheet_name] = df.sort_values(by='权重', ascending=False)
        
        # 刷新显示
        self.refresh_treeview(sheet_name)

    def refresh_treeview(self, sheet_name):
        """刷新树形视图显示"""
        tree = self.treeviews[sheet_name]
        tree.delete(*tree.get_children())
        
        for _, row in self.data[sheet_name].iterrows():
            tree.insert("", tk.END, values=tuple(row))

    def create_sheet_tab(self, sheet_name):
        """创建单个工作表页签"""
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text=sheet_name)
        
        # 创建Treeview
        tree = ttk.Treeview(frame, columns=("事件", "权重"), show='headings')
        tree.heading("事件", text="事件")
        tree.heading("权重", text="权重")
        
        # 添加滚动条
        scrollbar = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        tree.configure(yscrollcommand=scrollbar.set)
        
        # 布局
        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # 绑定双击事件
        tree.bind("<Double-1>", self.on_cell_edit)
        
        # 填充数据
        for _, row in self.data[sheet_name].iterrows():
            tree.insert("", tk.END, values=tuple(row))
        
        self.treeviews[sheet_name] = tree

    def on_cell_edit(self, event):
        """修复索引越界问题的编辑方法"""
        tree = event.widget
        
        # 获取点击位置信息
        region = tree.identify_region(event.x, event.y)
        if region not in ("cell", "tree"):
            return  # 忽略非单元格区域的点击
        
        # 获取当前选中项
        item = tree.identify_row(event.y)
        if not item:
            return  # 没有选中任何项
        
        # 获取列信息
        column = tree.identify_column(event.x)
        col_idx = int(column[1:]) - 1
        
        # 安全获取当前值
        try:
            current_values = list(tree.item(item, "values"))
            if len(current_values) != 2:
                raise ValueError("数据格式错误")
        except Exception as e:
            messagebox.showerror("错误", f"数据异常：{str(e)}")
            return
        
        # 创建编辑框
        x, y, width, height = tree.bbox(item, column)
        entry = ttk.Entry(tree)
        entry.place(x=x, y=y, width=width, height=height)
        
        # 设置初始值并绑定事件
        entry.insert(0, str(current_values[col_idx]))
        entry.select_range(0, tk.END)
        
        def commit_edit():
            new_value = entry.get()
            
            # 数据校验
            validation_passed = True
            if col_idx == 0:  # 事件列
                if not new_value.strip():
                    messagebox.showerror("错误", "事件内容不能为空")
                    validation_passed = False
            else:  # 权重列
                try:
                    float_value = float(new_value)
                    if float_value < 0:
                        raise ValueError
                    new_value = f"{float_value:.2f}"  # 格式化保留两位小数
                except ValueError:
                    messagebox.showerror("错误", "请输入有效的非负数字")
                    validation_passed = False
            
            if validation_passed:
                # 更新数据
                current_values[col_idx] = new_value
                tree.item(item, values=current_values)
                
                # 更新内存数据
                sheet_name = self.notebook.tab(self.notebook.select(), "text")
                self.update_sheet_data(sheet_name, tree)
            
            entry.destroy()
        
        entry.bind("<FocusOut>", lambda e: commit_edit())
        entry.bind("<Return>", lambda e: commit_edit())
        entry.focus_set()



    def show_add_dialog(self):
        """显示添加记录对话框"""
        dialog = tk.Toplevel(self.master)
        dialog.title("添加新记录")
        
        # 表单元素
        ttk.Label(dialog, text="事件：").grid(row=0, column=0, padx=5, pady=5)
        event_entry = ttk.Entry(dialog)
        event_entry.grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(dialog, text="权重：").grid(row=1, column=0, padx=5, pady=5)
        weight_entry = ttk.Entry(dialog)
        weight_entry.grid(row=1, column=1, padx=5, pady=5)
        
        def add_record():
            event = event_entry.get().strip()
            weight = weight_entry.get()
            
            # 输入校验
            if not event:
                messagebox.showerror("错误", "事件不能为空")
                return
            try:
                weight = float(weight)
                if weight < 0:
                    raise ValueError
            except ValueError:
                messagebox.showerror("错误", "请输入有效的非负数值")
                return
            
            # 获取当前页签
            sheet_name = self.notebook.tab(self.notebook.select(), "text")
            tree = self.treeviews[sheet_name]
            
            # 添加记录
            tree.insert("", tk.END, values=(event, weight))
            self.update_sheet_data(sheet_name, tree)
            dialog.destroy()
        
        ttk.Button(dialog, text="添加", command=add_record).grid(row=2, columnspan=2, pady=5)

    def delete_selected(self):
        """删除选中记录"""
        sheet_name = self.notebook.tab(self.notebook.select(), "text")
        tree = self.treeviews[sheet_name]
        
        selected = tree.selection()
        if not selected:
            messagebox.showwarning("警告", "请先选择要删除的记录")
            return
        
        for item in selected:
            tree.delete(item)
        
        self.update_sheet_data(sheet_name, tree)

    def save_data(self):
        """保存数据到Excel文件"""
        try:
            with pd.ExcelWriter(resource_path('events.xlsx'), engine='openpyxl') as writer:
                for sheet_name, df in self.data.items():
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            messagebox.showinfo("成功", "数据保存成功")
            self.keep_window_top()
        except Exception as e:
            messagebox.showerror("错误", f"保存失败：{str(e)}")
            self.keep_window_top()
    
    def keep_window_top(self):
            """保持窗口置顶状态的通用方法"""
            self.master.lift()  # 提升窗口到最前
            self.master.attributes('-topmost', True)  # 强制置顶
            self.master.after(100, lambda: self.master.attributes('-topmost', False))  # 恢复正常状态


class RedirectText:
    """重定向标准输出到文本组件"""
    def __init__(self, text_widget):
        self.text_widget = text_widget

    def write(self, string):
        self.text_widget.insert(tk.END, string)
        self.text_widget.see(tk.END)  # 自动滚动到底部

    def flush(self):
        pass

class FinanceApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("电访记录生成器")
        self.geometry("600x300")
        self.resizable(True, True)
        self.record_counts = []  # 存储记录数量的数组
        self.editor_window = None  # 新增窗口引用
        # self.create_ui()
        
        # 创建界面组件
        self.create_widgets()
        
        # 重定向标准输出
        sys.stdout = RedirectText(self.output_text)

    def create_widgets(self):
        # 主容器
        main_frame = ttk.Frame(self, padding=10)
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 输入区
        input_frame = ttk.LabelFrame(main_frame, text="记录生成设置", padding=10)
        input_frame.pack(fill=tk.X, pady=5)

        labels = ["存款记录数：", "理财记录数：", "贷款记录数："]
        self.entries = []
        for i, text in enumerate(labels):
            row = ttk.Frame(input_frame)
            row.pack(fill=tk.X, pady=3)
            ttk.Label(row, text=text, width=12).pack(side=tk.LEFT)
            entry = ttk.Entry(row)
            entry.pack(side=tk.RIGHT, expand=True, fill=tk.X)
            self.entries.append(entry)

        # 按钮区
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=5)
        
        ttk.Button(btn_frame, text="事件列表", command=self.show_editor).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="生成记录", command=self.create_fakedata).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="打开记录", command=self.open_excel).pack(side=tk.RIGHT, padx=5)

        # 输出区
        output_frame = ttk.LabelFrame(main_frame, text="执行日志", padding=6)
        output_frame.pack(fill=tk.BOTH, expand=True)

        self.output_text = tk.Text(output_frame, wrap=tk.WORD, height=6)
        scrollbar = ttk.Scrollbar(output_frame, command=self.output_text.yview)
        self.output_text.configure(yscrollcommand=scrollbar.set)

        self.output_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 编辑器容器
        self.editor = None

    def show_editor(self):
        """显示数据编辑器"""
        # 关闭已存在的编辑器窗口
        if self.editor_window and self.editor_window.winfo_exists():
            self.editor_window.destroy()
        
        # 创建新窗口
        self.editor_window = tk.Toplevel(self)
        self.editor_window.title("事件数据编辑")
        self.editor_window.geometry("600x400")

        # 设置窗口关系
        self.editor_window.transient(self)  # 设置为父窗口的临时窗口
        self.editor_window.grab_set()       # 独占焦点
        
        # 绑定关闭事件
        self.editor_window.protocol("WM_DELETE_WINDOW", self.close_editor)
        
        # 初始化编辑器
        ExcelEditor(self.editor_window)

    def close_editor(self):
        """安全关闭编辑器"""
        if messagebox.askokcancel("关闭", "确定要关闭编辑器吗？未保存的修改将会丢失"):
            self.editor_window.destroy()
            self.editor_window = None

    def create_fakedata(self):
        """执行生成操作"""
        self.record_counts = []
        for entry in self.entries:
            try:
                # 获取输入并去除空格
                input_value = entry.get().strip()
                # 空值处理为0
                value = int(input_value) if input_value else 0
            except ValueError:
                value = 0
            self.record_counts.append(value)
        def generate_fake_data(num_entries):
            # 生成符合规则的虚拟身份证号码
            def generate_id_number(region_code, gender,num = 1):
                """
                参数：
                region_code: 前6位地区代码（字符串）
                gender: 性别（'男'/'女' 或 1/0）
                num: 生成数量
                """
                def calc_check_code(first_17):
                    # 计算校验码
                    weight = [7,9,10,5,8,4,2,1,6,3,7,9,10,5,8,4,2]
                    check_code_map = {0:'1',1:'0',2:'X',3:'9',4:'8',5:'7',6:'6',7:'5',8:'4',9:'3',10:'2'}
                    total = sum(int(n) * weight[i] for i, n in enumerate(first_17))
                    return check_code_map[total % 11]

                # 验证参数
                if len(str(region_code)) != 6 or not str(region_code).isdigit():
                    raise ValueError("无效的地区代码")          
                if isinstance(gender, str):
                    gender = 1 if gender in ['男', 'male'] else 0
                elif gender not in [0, 1]:
                    raise ValueError("性别参数错误")

                # 生成年龄的正态分布（均值52.5，标准差13.75）
                ages = np.clip(np.random.normal(loc=52.5, scale=13.75, size=num), 25, 80)       
                id_numbers = []
                for age in ages:
                    # 生成出生日期
                    birth_date = datetime.now() - timedelta(days=age*365 + random.randint(0,364))           
                    # 生成顺序码（3位）
                    seq_code = f"{random.randint(0,99):02d}"
                    seq_code += str(random.randrange(1 if gender else 0, 10, 2))           
                    # 组合前17位
                    first_17 = f"{region_code}{birth_date.strftime('%Y%m%d')}{seq_code}"           
                    # 计算校验码
                    check_code = calc_check_code(first_17)           
                    id_numbers.append(first_17 + check_code)
                return id_numbers if num > 1 else id_numbers[0]
            fake_data = []
            
            for _ in range(num_entries):
                # 根据权重随机选择一个市
                chosen_city = np.random.choice(cities, p = list(map(lambda x: x / sum(city_weights), city_weights)))  
                # 筛选出包含所选市的区域名称的行
                chosen_city_df = df[df['region_name'].str.contains(chosen_city)]
                # 如果找到了包含所选市的行，则从中随机选择一行输出
                if not chosen_city_df.empty:
                    chosen_row = chosen_city_df.sample()
                    region_code = chosen_row['region_code'].values[0]
                    #region_name = chosen_row['region_name'].values[0]
                else:
                    print(f"No regions found for the city: {chosen_city}")

                # 筛选出包含所选城市的手机号代码的行
                chosen_phone_dp = dp[dp['region_name'].str.contains(chosen_city)]
                # 如果找到了包含所选市的行，则从中随机选择一行输出
                if not chosen_phone_dp.empty:
                    chosen_row_phone = chosen_phone_dp.sample()
                    phone_code = chosen_row_phone['phone_code'].values[0]
                    phone = str(phone_code) + str(random.randint(1110, 9999))
                else:
                    print(f"No phone_code found for the city: {chosen_city}")   
                
                # 生成名字、性别和年龄
                def generate_name(): 
                    if random.random() < 0.8: # 80%三字 
                        while True: 
                            last = fake.last_name() 
                            if len(last) == 1: # 需要双字名 
                                first = fake.first_name() 
                                if len(first) == 2: 
                                    return last + first 
                            elif len(last) == 2: # 需要单字名 
                                first = fake.first_name() 
                                if len(first) == 1: 
                                    return last + first 
                    else: # 20%二字 
                        while True: 
                            last = fake.last_name() 
                            if len(last) == 1: 
                                first = fake.first_name() 
                                if len(first) == 1: 
                                    return last + first
                name = generate_name()
                gender = 'male' if np.random.rand() < 0.4 else 'female'  # 男女比例2:3
                age = np.random.normal(52.5, 10)
                age = int(max(25, min(age, 80)))  # 限定年龄在25到80岁之间     
                # 生成身份证号
                id_number = generate_id_number(region_code, gender)
                fake_data.append({
                    '姓名': name,
                    '电话': phone,
                    '基础信息': id_number,
                })
            
            return fake_data

        # 生成虚假数据
        num_entries = self.record_counts[0] + self.record_counts[1] + self.record_counts[2]
        # 读取CSV文件
        df = pd.read_csv(resource_path('area.csv'), header=None, names=['region_code', 'region_name']) # 地区代码
        dp = pd.read_csv(resource_path('phone.csv'), header=None, names=['phone_code', 'region_name','city_code','operator','type']) # 电话代码

        # 设定每个市出现的概率
        cities = ["太原","晋中","大同","运城","忻州","吕梁","临汾","晋城","朔州","长治","阳泉"]
        city_weights = [90,1,1,1,1,1,1,1,1,1,1]

        fake = Faker(locale='zh_CN')
        fake_data_list = generate_fake_data(num_entries)
        # 存储最终选择的事件
        selected_events = []
        # 定义三个表页面的名称
        sheet_names = ['存款', '理财', '贷款']
        # 定义每张表需要随机选择的事件数量
        num_events_to_select = [self.record_counts[0], self.record_counts[1], self.record_counts[2]]  # 例如：从Sheet1选2个，Sheet2选3个，Sheet3选1个
        # 遍历每张表
        for i, sheet_name in enumerate(sheet_names):
            # 读取当前表
            de = pd.read_excel(resource_path('events.xlsx'), sheet_name=sheet_name)
            
            # 提取事件和概率列
            events = de.iloc[:, 0].values  # 第一列：事件
            probabilities = de.iloc[:, 1].values  # 第二列：概率
            
            # 将概率归一化（确保总和为1）
            probabilities = probabilities / np.sum(probabilities)
            
            # 按照概率随机选择指定数量的事件
            selected = np.random.choice(events, size=num_events_to_select[i], p=probabilities, replace=True)
            
            # 将选择的事件添加到结果中
            selected_events.extend(selected)
        for i in range(len(fake_data_list)):
            fake_data_list[i]['跟进记录'] = selected_events[i]
        # 加载已有文件
        wb = load_workbook(resource_path("电访记录表.xlsx"))
        ws = wb.active


        # 打印虚假数据
        for row_num, person in enumerate(fake_data_list, start=2):  # 从第2行开始
                # B列（第2列）写姓名
                ws[f'B{row_num}'] = person.get('姓名', '')
                
                # C列（第3列）写手机号
                ws[f'C{row_num}'] = person.get('电话', '')
                
                # D列（第4列）写身份证号
                ws[f'D{row_num}'] = person.get('基础信息', '')
                
                # F列（第6列）写年龄
                ws[f'G{row_num}'] = person.get('跟进记录', '')
                print(person.get('姓名', ''),person.get('电话', ''),person.get('基础信息', ''),person.get('跟进记录', ''))

        # 保存文件        
        try:
            wb.save(resource_path("电访记录表.xlsx"))
            print("数据已成功写入!")
                
        except Exception as e:
            print(f"生成失败：{str(e)}")           
            print(f"文件被占用，请关闭'电访记录表.xlsx'文件后重试")   


    def open_excel(self):
        """打开Excel文件"""
        file_path = resource_path("电访记录表.xlsx")
        
        if os.path.exists(file_path):
            try:
                os.startfile(file_path)
                print(f"\n已打开文件：{file_path}")
            except Exception as e:
                messagebox.showerror("打开失败", f"无法打开文件：{str(e)}")
        else:
            messagebox.showerror("文件不存在", f"未找到文件：{file_path}")


def resource_path(relative_path):
    """ 获取打包后资源的绝对路径 """
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)


if __name__ == "__main__":
    app = FinanceApp()
    app.mainloop()
